import time
from protectCalc import *
from openpyxl import *
#from g13Nodes import *
from cud1 import *

class CT(object):
    '''
        Classe que modela o TC, ***
    '''
    def __init__(self,
                name,
                Ipn = 500,
                Isn = 5,
                accuracyClass = None,
                Isat = None):
        self.name = name                        # Identificacao do TC
        self.Ipn = Ipn                          # Corrente nominal do primario do TC
        self.Isn = Isn                          # Corrente nominal do secundario do TC
        self.rtc = float(Ipn/Isn)               # Relacao de transformacao do TC
        self.current = 0                        # Leitura atual de corrente do primario TC
        self.accuracyClass = accuracyClass      # Classe de exatidao do TC (valores tipicos para TC 
                                                # de protecao: 2.5/5.0/10.0)

        if Isat is None:                        # Corrente do primario que leva o TC ao estado de saturacao
            self.Isat = 20*Ipn                  # Caso nao seja passada na criacao do objeto, assume 20*Ipn
        else:
            self.Isat = Isat

        self.isSaturated = False                # Flag que indica se o TC entrou em estado de saturacao

    def offSetCurrent(self,Ip):
        '''
            Funcao que retorna a corrente do secundario TC quando o primario
            e percorrido por uma corrente Ip.
        '''

        if self.Isat<Ip or self.isSaturated:        # Testa se o TC ficou ou esta saturado
            self.setSaturation()
            return float(self.Isat)/self.rtc
        else:
            return float(Ip)/self.rtc

    def setPrimaryCurrent(self,Ip):
        self.current=Ip
    
    def setSaturation(self,saturation=True):
        '''
            Funcao que seta o estado do TC para saturado (saturation=True) ou
            funcional (Saturation=False). Por padrao, se nenhum parametro eh 
            passado, o TC sera saturado.
        '''

        if saturation:
            print (self.name+" saturou. Prever manutencao.")
            self.isSaturated = True             # Em caso de saturacao, eh exibido um alerta de manutencao
        else:                                   # do TC. Eh necessario setar o parametro isSaturated = False
            self.isSaturated = False            # para que o TC tenha sua funcionalidade reestabelecida


    def autoSetting(self, Inom_alim, Iccmax):
        Ipnorma = [10, 20, 50, 100, 200, 400, 800, 1000, 2000, 5000]
        FT = 1.2 # Fator térmico, valor geral
        FScor = 20 # Fator de sobrecorrente, valor geral (sem saturar)
        limi = 2*Inom_alim/FT
        limf = Iccmax/FScor

        if limf < limi :
            limi = 1.5*Inom_alim/FT
            if limf < limi:
                print ("Impossível selecionar TC, rever dados.")
        
        for i in Ipnorma:
            if (limi<i and i<limf):
                self.Ipn = i
 
                        
class IED(object):
    '''
        Classe que modela o IED, ***
    '''
    def __init__(self,                      
                name,
                nmax,
                functions=None,
                ct=None,
                activeGroup=None,
                limits=None):

        # parametros do IED, passados na criacao do objeto

        self.name = name
        self.functions = functions              # dicionario com as funcoes habilitadas no ied **>implementar logica
        self.nmax = nmax                        # numero maximo de grupos de ajustes
        
        # Laco que associa os TCs de fase e neutro ao IED
        if ct is None:
            self.ct=[CT("TC A do "+self.name)]
            self.ct.append(CT("TC B do "+self.name))
            self.ct.append(CT("TC C do "+self.name))
            self.ct.append(CT("TC N do "+self.name))
        else:
            self.ct = ct        

        self.activeGroup = activeGroup          # objeto da classe AdjustGroup contendo o grupo de
                                                # ajustes ativo do IED
        self.BRKF = False                       # flag para habilitar/desabilitar falha de disjuntor
        self.function_sens = None               # flag que indica a funcao de protecao indicada no 
                                                # painel do ied
        self.current = float()                  # inicializa a corrente do ied como zero

        self.trip = False

        if limits is not None:                  # limites permitidos pelo rele

            self.iaj51_min = limits['I51MIN']
            self.iaj51_max = limits['I51MAX']

            self.iaj50_min = limits['I50MIN']
            self.iaj50_max = limits['I50MAX']

            self.dial_min = limits['DIAL_MIN']
            self.dial_max = limits['DIAL_MAX']

        # loop cria variavel que contera todos os grupos de ajustes cadastrados

        self.groups = dict()

        for i in range(nmax):

            tag = str(i+1)
            self.groups[tag] = None

    def sendTrip(self):
        currents = []
        for t in self.ct:
            currents.append(t.current)
    
        if self.activeGroup.tripF50(max(currents)):
            self.function_sens = '50'
            self.trip = True
        elif self.activeGroup.tripF50N(currents[3]):
            self.function_sens = '50N'
            self.trip = True
        elif self.activeGroup.tripF51(max(currents)):
            time.sleep(self.activeGroup.delayF51(max(currents)))
            self.function_sens = '51'
            self.trip = True
        elif self.activeGroup.tripF51N(currents[3]):
            time.sleep(self.activeGroup.delayF51N(currents[3]))
            self.function_sens = '51N'
            self.trip = True

    def reset(self):
        self.trip = False
        self.current = 0
    
    def calculateCurveP(self,
                    iscmax=None,
                    delay=0.3):
        '''
                Método que calcula a curva de atuação (fase) do IED, para
                então ser plotada por um dos métodos de plotagem, recebe como parâmetros:
                Iscmax: Maior corrente de curto circuito a ser plotada. Quando
                    não passada, assume valor de 150%' da corrente de pick up
                    instantânea.
                delay: "Distância" do ultimo elemento de proteção do ramo:
                    0 - para o elemento mais a jusante.
                    1 - para o elemento imediatamente a montante do elemento 0
                    2 - para o elemento imediatamente a montante do elemento 1
                    E assim sucessivamente.
                    obs: se for float, passa automaticamente o tempo
                
                Retorna como parâmetros:
                    isc - lista de correntes de curto (eixo x)
                    opTime5051 - lista de tempos de atuação (eixo y)
        '''
        #distance: "Distância" do ultimo elemento de proteção do ramo:
                    #0 - para o elemento mais a jusante.
                    #1 - para o elemento imediatamente a montante do elemento 0
                    #2 - para o elemento imediatamente a montante do elemento 1
                    #E assim sucessivamente.

        import numpy as np

        if(iscmax==None):
            iscmax=int(self.activeGroup.ipk50*1.5)

        isc = []
        opTime5051 = []
        #dist=delay

        #if isinstance(delay, int):
            #self.activeGroup.delay50=delay*0.03

        isc=range(int(self.activeGroup.ipk51*1.4),iscmax)
        for i in isc:
            if self.activeGroup.ipk51 >= i:
                #opTime5051.append(self.activeGroup.ipk51+1)
                pass
            elif self.activeGroup.ipk50 > i:
                opTime5051.append(self.activeGroup.delayF51(i))
            else:
                opTime5051.append(self.activeGroup.delay50)
        return isc,opTime5051

    def calculateCurveN(self,
                    iscmax=None,
                    delay=0.3):
        '''
                Método que calcula a curva de atuação (neutro) do IED, para
                então ser plotada por um dos métodos de plotagem, recebe como parâmetros:
                Iscmax: Maior corrente de curto circuito a ser plotada. Quando
                    não passada, assume valor de 150%' da corrente de pick up
                    instantânea.
                delay: "Distância" do ultimo elemento de proteção do ramo:
                    0 - para o elemento mais a jusante.
                    1 - para o elemento imediatamente a montante do elemento 0
                    2 - para o elemento imediatamente a montante do elemento 1
                    E assim sucessivamente.
                    obs: se for float, passa automaticamente o tempo
                
                Retorna como parâmetros:
                    iscN - lista de correntes de curto (eixo x)
                    opTime5051N - lista de tempos de atuação (eixo y)
        '''

        import numpy as np

        if(iscmax==None):
            iscmax=int(self.activeGroup.ipk50N*1.5)

        iscN = []
        opTime5051N = []
        #dist=delay

        #if isinstance(delay, int):
            #self.activeGroup.delay50N=delay*0.03

        iscN=range(int(self.activeGroup.ipk51N*1.4),iscmax)
        for i in iscN:
            if self.activeGroup.ipk51N >= i:
                #opTime5051N.append(self.activeGroup.ipk51N+1)
                pass
            elif self.activeGroup.ipk50N > i:
                opTime5051N.append(self.activeGroup.delayF51N(i))
            else:
                opTime5051N.append(self.activeGroup.delay50N)
        return iscN,opTime5051N


    def plotCurve(self,
                  iscmax=None,
                  distance=0,
                  PN='P'):
        '''
            Método que plota a curva do IED, recebe como parâmetros:
                Iscmax: Maior corrente de curto circuito a ser plotada. Quando
                    não passada, assume valor de 150%' da corrente de pick up
                    instantânea.
                distance: "Distância" do ultimo elemento de proteção do ramo:
                    0 - para o elemento mais a jusante.
                    1 - para o elemento imediatamente a montante do elemento 0
                    2 - para o elemento imediatamente a montante do elemento 1
                    E assim sucessivamente.
                PN: Define o tipo de curva a ser plotada:
                    P - phase/fase
                    N - neutral/neutro 
        '''

        import matplotlib.pyplot as plt
        
        if PN=='P':
            color='r'
            i,t=self.calculateCurveP(iscmax,distance)
            plt.plot(i,t,color)
        elif PN=='N':
            color='b--'
            i,t=self.calculateCurveN(iscmax,distance)
            plt.plot(i,t,color)
        else:
            color='r'
            i,t=self.calculateCurveP(iscmax,distance)
            plt.plot(i,t,color)
            color='b--'
            i,t=self.calculateCurveN(iscmax,distance)
            plt.plot(i,t,color)

        plt.show()

    def autoSettingCoord (self, KF, Inom_alim, I2fmin, Iftmin):
        FS = 1.4 #Fator de segurança, em média 1.4
        FI = 1.5 #Fator de início da curva
        Kenerg = 3 #Constante de energização 3 a 8, inrush????
        Icarga = 2*Inom_alim
        
        '''
        for i in self.ct:
            i.autoSetting(Inom_alim,20*Inom_alim)

        RTC51 = self.ct[0].Ipn/self.ct[0].Isn
        RTC51N = self.ct[3].Ipn/self.ct[3].Isn
        '''
        
        coord = True

        limi51 = Icarga*KF#/RTC51
        limf51 = I2fmin/(FS*FI)#*RTC51)

        if limf51 < limi51 and coord:
            Icarga = 1.5*Inom_alim
            limi51 = Icarga*KF#/RTC51
            if limf10 < limi51:
                print("Impossível coordenar IED, rever dados.")
                coord = False

        limi51N = 0.3*Icarga#/RTC51N
        limf51N = Iftmin/(FS*FI)#*RTC51N)

        if limf51N < limi51N and coord:
            print("Impossível coordenar IED, rever dados.")
            coord = False

        limi50 = Kenerg*Icarga#/RTC51
        limf50 = I2fmin#/(RTC51)

        if limf50 < limi50 and coord:
            print("Impossível coordenar IED, rever dados.")
            coord = False


        limi50N = 0.3*Icarga#/RTC51N
        limf50N = Iftmin#/(RTC51N)

        if limf50N < limi50N and coord:
            print ("Impossível coordenar IED, rever dados.")
            coord = False

        self.limi50 = limi50
        self.limf50 = limf50
        self.limi51 = limi51
        self.limf51 = limf51
        self.limi50N = limi50N
        self.limf50N = limf50N
        self.limi51N = limi51N
        self.limf51N = limf51N

        if coord:
            self.activeGroup.ipk51 = (limf51+limi51)/2
            self.activeGroup.ipk51N = (limf51N+limi51N)/2
            self.activeGroup.ipk50 = limi50+(limf50-limi50)/3
            self.activeGroup.ipk50N = (limf50N+limi50N)/2 #especial


    def autoSettingSelect (self,KF,Inom_alim,I2fmin,I2fmin_sec,Iftmin,Iftmax):
        FS = 1.4 #Fator de segurança, em média 1.4
        FI = 1.5 #Fator de início da curva
        Kenerg = 3 #Constante de energização 3 a 8, inrush????

        '''
        for i in self.ct:
            i.autoSetting(Inom_alim,20*Inom_alim)
        RTC51 = self.ct[0].Ipn/self.ct[0].Isn
        RTC51N = self.ct[3].Ipn/self.ct[3].Isn
        '''

        Icarga = 2*Inom_alim

        select = True

        limi51 = Icarga*KF#/RTC51
        limf51 = I2fmin/(FS*FI)#*RTC51)

        if limf51 < limi51 :
            Icarga = 1.5*Inom_alim
            limi51 = Icarga*KF#/RTC51
            if limf10 < limi51:
                print("Impossível selecionar IED, rever dados.")
                select = False

        limi51N = 0.3*Icarga#/RTC51N
        limf51N = Iftmin/(FS*FI)#*RTC51N)

        if limf51N < limi51N and select:
            print("Impossível selecionar IED, rever dados.")
            select = False

        self.activeGroup.ipk51N = limi51N

        limi50 = Kenerg*Icarga#/RTC51
        limf50 = I2fmin_sec#/(RTC51) #Ponto B, página 62

        if limf50 < limi50 and select:
            print("Impossível selecionar IED, rever dados.")
            select = False

        limi50N = Iftmax#/(RTC51N)

        self.limi50 = limi50
        self.limf50 = limf50
        self.limi51 = limi51
        self.limf51 = limf51
        self.limi50N = limi50N
        #self.limf50N = limf50N
        self.limi51N = limi51N
        self.limf51N = limf51N

        if select:
            self.activeGroup.ipk51 = (limf51+limi51)/2
            self.activeGroup.ipk51N = (limf51N+limi51N)/2
            self.activeGroup.ipk50 = limf50-1
            self.activeGroup.ipk50N = limi50N+1 #especial

class FuseSwitch(object):
    '''
        Classe que modela a chave fusível, ***
    '''
    def __init__(self,               
                name,
                link,
                In=0,
                IscS=0,
                IscA=0,
                voltage=0):
        self.name = name
        self.In = In
        self.IscS = IscS 
        self.IscA = IscA
        self.voltage = voltage
        self.link = link


class FuseLink(object):
    '''
        Classe que modela a chave fusível, ***
    '''
    def __init__(self,
                curveType,
                In):
        self.curveType = curveType
        self.In = In    
        self.standardIn()
        self.loadCurve()

    def loadCurve(self):
        self.curve = [[],[],[],[]]
        curves = load_workbook("CURVAS.xlsx")
        planCurves = curves.active
        for i in range(200):
            if planCurves.cell(column=i+1, row=1).value == str(str(int(self.In))+self.curveType+"Imin"):
                for j in range(100):
                    if planCurves.cell(column=i+1, row=j+2).value != None and planCurves.cell(column=i+1, row=j+2).value != "":
                        self.curve[0].append(float(planCurves.cell(column=i+1, row=j+2).value))
                        self.curve[1].append(float(planCurves.cell(column=i+2, row=j+2).value))
            if planCurves.cell(column=i+1, row=1).value == str(str(int(self.In))+self.curveType+"Imax"):
                for j in range(100):
                    if planCurves.cell(column=i+1, row=j+2).value != None and planCurves.cell(column=i+1, row=j+2).value != "":
                        self.curve[2].append(float(planCurves.cell(column=i+1, row=j+2).value))
                        self.curve[3].append(float(planCurves.cell(column=i+2, row=j+2).value))

    def standardIn(self):
        curves = load_workbook("CURVAS.xlsx")
        planCurves = curves.active
        standard = []
        for i in range(200):
            Ist = planCurves.cell(column=i+1, row=1).value
            if Ist != None and Ist != "":               
                if Ist[-5:] == str(self.curveType+"Tmin"):
                    standard.append(float(Ist[:-5]))
        for i in standard:
            if i >= self.In:
                self.In = i
                #print("Mudou In "+str(self.In))
                break

    def setCurveType(self,x):
        if (x == "K" or x == "H" or x == "T"):
            self.curveType = x
            #print("Mudou curva "+self.curveType)

    def autoSetting(self, KF, Inom_alim, Iftmin):
        limi = KF*2*Inom_alim
        limf = (0.25)*Iftmin

        print(limi)
        print(limf)

        flag = True
        if limf < limi :
            limi = KF*1.5*Inom_alim
            if limf < limi:
                print("Impossível selecionar Chave-Fusível, rever dados.")
                flag = False

        if flag:
            self.In = limi
            self.standardIn()

        # K - Elo rápido, fusão entre 6 a 8,1 vezes a self.In (disp protetor)
        # T - Elo lento, fusão entre 10 a 13 vezes a self.In (disp protegido)
        # H - Elo para cargas com in-rush 


class Breaker(object):
    '''
        Classe que modela o disjuntor, não implementada.
    '''
    def __init__(self,
                name,
                opStatus='NC',
                relay=None):
        
        self.name = name                        # Nome do disjuntor
        self.relay = relay                      # Associacao do rele ao disjuntor
        if opState is 'NO':                     # Define o estado inicial do disjuntor pelo seu estado 
            self.status = 'open'                # normal de operacao: open quando 'NO'(normalmente aberto) 
        else:                                   #                     closed nos demais casos.
            self.status = 'closed'              # Por padrao, opState recebe 'NC'
    
    def isOpen(self):
        if self.status == 'open':
            return True
        else:
            return False

    def setOpen(self):
        self.state = 'open'

    def setClosed(self):
        self.state = 'closed'           


class AdjustGroup(object):
    '''
        Classe que modela um grupo de ajuste: cada objeto criado armazena
        as correntes de pickup das funcoes 50/51 e 50/51N, os tipos de
        curva e os valores de dial.
    '''
    def __init__(self,ipk50,ipk51,curveP,dialP,ipk50N,ipk51N,curveN,dialN):
        
        # ajustes de fase

        self.ipk50 = ipk50                      # ajuste da corrente de pickup da funcao 50

        self.ipk51 = ipk51                      # ajuste da corrente de pickup da funcao 51
        self.curveP = curveP                    # tipo ('NI','MI' ou 'EI') de curva de atuacao da funcao 50 
        self.dialP = dialP                      # dial da curva de atuacao da funcao 50 de fase
        self.delay50 = 0.4

        # ajustes de neutro

        self.ipk50N = ipk50N                    # ajuste da corrente de pickup da funcao 50N
        
        self.ipk51N = ipk51N                    # ajuste da corrente de pickup da funcao 51N
        self.curveN = curveN                    # tipo ('NI','MI' ou 'EI') de curva de atuacao da funcao 50N
        self.dialN = dialN                      # dial da curva de atuacao da funcao 50N
        self.delay50N = 0.4

    def returnParameters(self,curve):
        '''
            Funcao que retorna os parametros alfa e beta de acordo com o tipo
            de curva passada como parametro:
                'NI'/'SI': Normalmente Inversa
                'MI'/'VI': Muito Inversa
                'EI': Extremamente Inversa
        '''

        if curve == 'NI' or curve == 'SI':

            beta = 0.14
            alpha = 0.02

        elif curve == 'MI' or curve == 'VI':

            beta = 13.5
            alpha = 1.0

        elif curve == 'EI':

            beta = 80
            alpha = 2.0

        return alpha,beta

    def tripF50(self,Iphase):
        if self.ipk50 > Iphase:
            return False
        else:
            return True

    def tripF51(self,Iphase):
        if self.ipk51 < Iphase:
            time.sleep(self.delayF51(Iphase))                      #desabilitado para teste
            return True
        else:
            return False    

    def delayF51(self,Iphase):
        a,b = self.returnParameters(self.curveP)
        return self.dialP*b/(((Iphase/self.ipk51)**a)-1)    

    def tripF50N(self,Ineutral):
        if self.ipk50N > Ineutral:
            return False
        else:
            return True

    def tripF51N(self,Ineutral):
        if self.ipk51N < Ineutral:
            time.sleep(self.delayF51N(Ineutral))                   #desabilitado para teste
            return True
        else:
            return False

    def delayF51N(self,Ineutral):
        a,b = self.returnParameters(self.curveP)
        return self.dialN*b/(((Ineutral/self.ipk51N)**a)-1)

    def calculateDialP(self,time_ref,i_ref,delay):
        a,b = self.returnParameters(self.curveP)
        self.dialP = ((time_ref+delay)/b)*(((i_ref/self.ipk51)**a)-1)

    def calculateDialN(self,time_ref,i_ref,delay):
        a,b = self.returnParameters(self.curveN)
        self.dialN = ((time_ref+delay)/b)*(((i_ref/self.ipk51N)**a)-1)

class PowerGrid(object):
    '''
        Classe que modela uma rede eletrica.
    '''
    def __init__(self,grid_data,switches,time=0.4):
        self.grid_data = grid_data
        self.ieds = switches
        self.switches = switches
        self.time = 0.4        

    def addIED(self, ied):
        self.ieds.append(ied)

    def addSwitch(self, switch):
        self.switchs.append(switch)

    def plotCurves(self,iscmax=None,ieds=None,PN='P'):
        '''
            Método que plota a curva dos IEDs da rede, recebe como parâmetros:
                Iscmax: Maior corrente de curto circuito a ser plotada. Quando
                    não passada, assume valor de 150%' da corrente de pick up
                    instantânea.
                ieds: lista de ieds do sistema, do mais a jusante para o mais
                    a jusante. Quando não passada a lista, adota-se a lista dos
                    IEDs formados na criação da rede.
        '''

        import matplotlib.pyplot as plt
        import numpy as np

        if ieds == None:
            ieds = self.ieds

        j=0
        for ied in ieds:
            if PN=='P':
                color='r'
                i,t=ied.calculateCurveP(iscmax,j)
                plt.plot(i,t,color)
            elif PN=='N':
                color='b--'
                i,t=ied.calculateCurveN(iscmax,j)
                plt.plot(i,t,color)
            else:
                color='r'
                i,t=ied.calculateCurveP(iscmax,j)
                plt.plot(i,t,color)
                color='b--'
                i,t=ied.calculateCurveN(iscmax,j)
                plt.plot(i,t,color)
            j+=1
        plt.show()

    #def autoSetting(self, delta, n):
        #KF=(1+(delta/100))^n

    def autoSettingCoord(self):
        
        for i in range(len(self.ieds)):
            if isinstance(self.ieds[i], FuseSwitch):
                name = self.ieds[i].name
                for j in self.grid_data.sections:
                    if 'S'+grid_elements.sections[j].switch.name == name:
                        Inom_alim = []
                        Iftmin = []

                        flag2=True
                        l=j
                        if l[2:3]=='F':
                            for w in grid_data.sections:
                                if j[2:4] == w[0:2] and flag2:
                                    l=w

                        self.grid_data.dist_grids['F0'].sections[l].n2._calc_currents()
                        for In in self.grid_data.dist_grids['F0'].sections[l].n2.i:
                            Inom_alim.append(np.absolute(In))
                        Inom_alim = max(Inom_alim)
                        #for Ift in mono_phase(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                        for Ift in min_mono_phase(self.grid_data.dist_grids['F0'],j[2:4],zt=0)[j[2:4]]:
                            Iftmin.append(np.absolute(Ift))
                        Iftmin = max(Iftmin)
                self.ieds[i].link.autoSetting(1,Inom_alim,Iftmin)

            if isinstance(self.ieds[i], IED):
                name = self.ieds[i].name
                for k in self.grid_data.sections:
                    if 'S'+grid_elements.sections[k].switch.name == name:
                        Inom_alim = []
                        Iftmin = []
                        Iftmax = []
                        I2f = []

                        flag2=True
                        j=k
                        for w in grid_data.sections:
                            if k[2:4] == w[0:2] and flag2:
                                j=w

                        flag2=True
                        l=k
                        if l[2:3]=='F':
                            for w in grid_data.sections:
                                if k[2:4] == w[0:2] and flag2:
                                    l=w

                        self.grid_data.dist_grids['F0'].sections[l].n2._calc_currents()
                        for In in self.grid_data.dist_grids['F0'].sections[l].n2.i:
                            Inom_alim.append(np.absolute(In))
                        Inom_alim = max(Inom_alim)[0]
                        for Ift in min_mono_phase(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                            Iftmin.append(np.absolute(Ift))
                        Iftmin = max(Iftmin)[0]
                        for If in min_mono_phase(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                            Iftmax.append(np.absolute(If))
                        Iftmax = max(Iftmax)[0]
                        for I2 in biphasic(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                            I2f.append(np.absolute(I2))
                        I2f = max(I2f)[0]
                self.ieds[i].autoSettingCoord(1,Inom_alim,I2f,Iftmin)

        for i in range(len(self.ieds)-1):
            if isinstance(self.ieds[len(self.ieds)-2-i], FuseSwitch):
                self.autoSwSwSelect(self.ieds[len(self.ieds)-1-i],self.ieds[len(self.ieds)-2-i])
            elif isinstance(self.ieds[len(self.ieds)-1-i], FuseSwitch):
                self.autoIEDSwCoord(self.ieds[len(self.ieds)-1-i],self.ieds[len(self.ieds)-2-i])                
            else:
                self.autoIEDIEDSelect(self.ieds[len(self.ieds)-1-i],self.ieds[len(self.ieds)-2-i])


    def autoSettingSelect(self):
        for i in range(len(self.ieds)):
            if isinstance(self.ieds[i], FuseSwitch):
                name = self.ieds[i].name
                for j in self.grid_data.sections:
                    if 'S'+grid_elements.sections[j].switch.name == name:
                        Inom_alim = []
                        Iftmin = []

                        flag2=True
                        l=j
                        if l[2:3]=='F':
                            for w in grid_data.sections:
                                if j[2:4] == w[0:2] and flag2:
                                    l=w


                        self.grid_data.dist_grids['F0'].sections[l].n2._calc_currents()
                        for In in self.grid_data.dist_grids['F0'].sections[l].n2.i:
                            print(np.absolute(In))
                            Inom_alim.append(np.absolute(In))
                        Inom_alim = max(Inom_alim)
                        #for Ift in mono_phase(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                        for Ift in min_mono_phase(self.grid_data.dist_grids['F0'],j[2:4],zt=0)[j[2:4]]:
                            print(np.absolute(Ift))
                            Iftmin.append(np.absolute(Ift))
                        Iftmin = max(Iftmin)
                self.ieds[i].link.autoSetting(1,Inom_alim,Iftmin)

            if isinstance(self.ieds[i], IED):
                name = self.ieds[i].name
                for k in self.grid_data.sections:
                    if 'S'+grid_elements.sections[k].switch.name == name:
                        Inom_alim = []
                        Iftmin = []
                        Iftmax = []
                        I2f = []
                        I2fsec = []

                        flag2=True
                        j=k
                        for w in grid_data.sections:
                            if k[2:4] == w[0:2] and flag2:
                                j=w

                        flag2=True
                        l=k
                        if l[2:3]=='F':
                            for w in grid_data.sections:
                                if k[2:4] == w[0:2] and flag2:
                                    l=w

                        self.grid_data.dist_grids['F0'].sections[l].n2._calc_currents()
                        for In in self.grid_data.dist_grids['F0'].sections[l].n2.i:
                            Inom_alim.append(np.absolute(In))
                        Inom_alim = max(Inom_alim)[0]
                        for Ift in min_mono_phase(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                            Iftmin.append(np.absolute(Ift))
                        Iftmin = max(Iftmin)[0]
                        for If in mono_phase(self.grid_data.dist_grids['F0'],k[2:4])[k[2:4]]:
                            Iftmax.append(np.absolute(If))
                        Iftmax = max(Iftmax)[0]
                        for I2 in biphasic(self.grid_data.dist_grids['F0'],j[2:4])[j[2:4]]:
                            I2f.append(np.absolute(I2))
                        I2f = max(I2f)[0]
                        for I2 in biphasic(self.grid_data.dist_grids['F0'],k[2:4])[k[2:4]]:
                            I2fsec.append(np.absolute(I2))
                        I2fsec = max(I2fsec)[0]
                self.ieds[i].autoSettingSelect(1,Inom_alim,I2f,I2fsec,Iftmin,Iftmax)
            
        for i in range(len(self.ieds)-1):
            #print("Entrou"+str(i))
            if isinstance(self.ieds[len(self.ieds)-2-i], FuseSwitch):
                self.autoSwSwSelect(self.ieds[len(self.ieds)-1-i],self.ieds[len(self.ieds)-2-i])
            elif isinstance(self.ieds[len(self.ieds)-1-i], FuseSwitch):
                self.autoIEDSwSelect(self.ieds[len(self.ieds)-1-i],self.ieds[len(self.ieds)-2-i])                
            else:
                self.autoIEDIEDSelect(self.ieds[len(self.ieds)-1-i],self.ieds[len(self.ieds)-2-i])
            

    def autoSwSwSelect(self, swp, swr):
        print("SwSw")
        swr.link.loadCurve()
        swp.link.loadCurve()

        select = False

        while not (select):
            '''
            flag=True
            i=0
            for j in (swr.link.curve[0]):
                i+=1
                if swp.link.curve[2][-1]<j and flag:
                    flag = False
                    print ("r:"+str(swr.link.curve[1][i]*0.75)+" e p:"+str(swp.link.curve[3][-1]))
                    if (swr.link.curve[1][i]*0.75>swp.link.curve[3][-1]): 
                        select = True
            flag=True
            i=0
            if select:
                for j in (swp.link.curve[2]):
                    i+=1
                    if swr.link.curve[0][0]<j and flag:
                        flag = False
                        print ("r:"+str(swr.link.curve[1][0]*0.75)+" e p:"+str(swp.link.curve[3][i]))
                        if (swr.link.curve[1][0]*0.75<swp.link.curve[3][i]):               
                            select = False
            '''

            med=round(len(swp.link.curve[0])/2)

            flag=True
            i=0
            for j in (swp.link.curve[0]):
                i+=1
                if swp.link.In*2>=j and flag:
                    med=i
                    flag = False

            flag=True
            i=0

            for j in (swr.link.curve[0]):
                print ("r:"+str(swr.link.curve[0][i])+" e p:"+str(swp.link.curve[2][med]))
                i+=1
                if swp.link.curve[2][med]>=j and flag:
                    flag = False
                    #print ("Entrou-> r:"+str(swr.link.curve[0][i])+" e p:"+str(swp.link.curve[2][med]))
                    if (swr.link.curve[1][i]*0.75>=swp.link.curve[3][med]): 
                        select = True


            if not (select):
                if swr.link.In<0.75*swp.link.In:
                    swr.link.In=0.75*swp.link.In
                if swr.link.curveType=="K":
                    swr.link.curveType="T"
                elif swr.link.curveType=="T":
                    swr.link.curveType="K"
                    swr.link.In+=1
                    swr.link.standardIn()
                else:
                    swr.link.curveType="K"
            swr.link.loadCurve()

            print(swr.link.In, swr.link.curveType)
        #print("End")


    def autoIEDIEDSelect(self, swp, swr):

        if self.ieds.index(swp)==len(self.ieds)-1:
            swp.activeGroup.delay50=self.time
            swp.activeGroup.delay50N=self.time
            swp.activeGroup.calculateDialP(swp.activeGroup.delay50,swp.activeGroup.ipk50,0)
            swp.activeGroup.calculateDialN(swp.activeGroup.delay50N,swp.activeGroup.ipk50N,0)
        
        swr.activeGroup.delay50=self.time#+swp.activeGroup.delay50
        swr.activeGroup.delay50N=self.time#+swp.activeGroup.delay50N
        swr.activeGroup.curveP=swp.activeGroup.curveP
        swr.activeGroup.curveN=swp.activeGroup.curveN

        tref=swp.activeGroup.delayF51(swp.activeGroup.ipk50)
        trefN=swp.activeGroup.delayF51N(swp.activeGroup.ipk50N)
        swr.activeGroup.calculateDialP(tref,swp.activeGroup.ipk50,self.time)
        swr.activeGroup.calculateDialN(trefN,swp.activeGroup.ipk50N,self.time)
        #swr.activeGroup.calculateDialN(trefN,swp.activeGroup.ipk50,self.time)

        #c,t = swr.calculateCurveP()
        #cn,tn = swr.calculateCurveN()
        #for i in c:     
        if swr.activeGroup.delayF51(swr.activeGroup.ipk50)<swr.activeGroup.delay50:
            swr.activeGroup.calculateDialP(swr.activeGroup.delay50,swr.activeGroup.ipk50,0)
        if swr.activeGroup.delayF51N(swr.activeGroup.ipk50N)<swr.activeGroup.delay50N:
            swr.activeGroup.calculateDialN(swr.activeGroup.delay50N,swr.activeGroup.ipk50N,0)

        

        print("IEDIED")
        pass

    def autoIEDSwSelect(self, swp, swr):
        #está acima

        flag=True
        i=0
        for j in (swp.link.curve[0]):
            i+=1
            print(swp.link.curve[0][i-1])
            if swr.activeGroup.ipk50>=j and flag:
                med=i
                flag = False

        swr.activeGroup.delay50=self.time#+swp.link.curve[1][med-1]
        swr.activeGroup.calculateDialP(swr.activeGroup.delay50,swp.link.curve[2][med],self.time)

        flag=True
        i=0
        for j in (swp.link.curve[0]):
            i+=1
            if swr.activeGroup.ipk50N>=j and flag:
                med=i
                flag = False

        swr.activeGroup.delay50N=self.time#+swp.link.curve[1][med-1]
        swr.activeGroup.calculateDialN(swr.activeGroup.delay50N,swp.link.curve[2][med],self.time)



        print("IEDSwSelect")
        pass
        
    def autoIEDSwCoord(self, swp, swr):
        #passa no meio
        #print("entrou!")
        flag=True
        i=0
        for j in (swp.link.curve[0]):
            i+=1
            print(swp.link.curve[0][i-1])
            if swr.activeGroup.ipk50>=j and flag:
                med=i
                flag = False

        swr.activeGroup.delay50=(self.time)#*(-1))+swp.link.curve[1][med-1]
        swr.activeGroup.calculateDialP(swp.link.curve[3][med],swp.link.curve[2][med],self.time)

        flag=True
        i=0
        for j in (swp.link.curve[0]):
            i+=1
            if swr.activeGroup.ipk50N>=j and flag:
                med=i
                flag = False

        swr.activeGroup.delay50N=(self.time)#*(-1))+swp.link.curve[1][med-1]
        swr.activeGroup.calculateDialN(swp.link.curve[3][med],swp.link.curve[2][med],self.time)




        print("IEDSwCoord")
        pass
        













        











